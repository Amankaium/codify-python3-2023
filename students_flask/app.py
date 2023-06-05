from flask import Flask, render_template, request
from openpyxl import load_workbook  # библиотека для работы с excel
from flask_sqlalchemy import SQLAlchemy  # библиотека для работы с БД
from random import randint

db = SQLAlchemy()
app = Flask(__name__)
# app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///project.db"
app.config["SQLALCHEMY_DATABASE_URI"] = "postgresql://postgres:postgres@localhost:5432/user_db"
db.init_app(app)


class UserTable(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String, unique=True, nullable=False)
    second_name = db.Column(db.String, nullable=True)
    year = db.Column(db.Integer, nullable=True)
    email = db.Column(db.String)
    is_active = db.Column(db.Boolean, default=False, nullable=False)


with app.app_context():
    db.create_all()


@app.route('/user-form', methods=['GET', 'POST'])
def user_add_database():
    if request.method == "POST":
        data = request.form  # словарь
        print(data)
        is_active = data.get('is_active') == "on"
        user = UserTable(
            username=data["username"],
            second_name=data["second_name"],
            year=data["year"],
            email=data["email"],
            is_active=is_active
        )
        db.session.add(user)  # INSERT INTO
        db.session.commit()  # COMMIT (END)
        return "Пользователь создан"
    return render_template('user_form.html')


@app.route("/create-user")
def user_create():
    num = randint(1, 10000)  # генерируется случайное число от 1 до 10000
    user = UserTable(
        username=f"user{num}",  # user472 , user4235
        second_name=f'Johnson{num}',  # Johnson472, Johnson4235
        year=randint(1985, 2010)
    ) # создаётся объект класса User
    db.session.add(user)  # INSERT INTO
    db.session.commit()  # COMMIT (END)
    return "Пользователь создан"


@app.route('/database')
def db_function():
    users = db.session.execute(db.select(UserTable).order_by(UserTable.username)).scalars()  # SELECT
    result = ''  # формируем список логинов пользователей
    for user in users:
        result += f"<div>{user.username} - {user.second_name} - {user.year} - {user.is_active}</div>"
    return f"{result}"


@app.route('/')
def hello_world():
    return """
        <nav>
            <ul>
                <li><a href="/">Домой</a></li>
                <li><a href="/my-code">Мой код</a></li>
                <li><a href="/add">Добавить студента</a></li>
            </ul>
        </nav>
        <div>
            Hello python-3 2023!
        </div>
        <a href="https://flask-russian-docs.readthedocs.io/ru/0.10.1/">
            Документация Flask на русском.
        </a>
    """


@app.route('/my-code')
def my_render():
    return render_template('index.html')


@app.route('/students')
def students():
    students_list = []
    my_file = load_workbook("python_students.xlsx")  # открываем файл через его путь
    page = my_file["Лист1"]  # обращаемся к нужной страница по имени
    students_cells = page["A"]  # берём список ячеек по имени колонны
    for cell in students_cells:
        students_list.append(cell.value)
    result = ""
    for index, student in enumerate(students_list):  # 0, "Elver"
        result += f"<div>{index+1}. {student}</div>"
    return result


@app.route('/marks')
def marks():
    excel_file = load_workbook('python_students.xlsx')
    page = excel_file["Лист1"]
    students_list = page["A"]  # [Cell, Cell, ...]
    marks_list = page['B']
    result = ''
    for i in range(len(students_list)):
        student = students_list[i].value
        mark = marks_list[i].value
        result = result + f'<div>{i+1}. {student} - {mark}</div>'
    return result


@app.route('/add', methods=["GET", "POST"])
def add_student():
    data = request.form  # dictionary
    # print(data)
    if data:
        new_student = data["student_name"]
        mark_1 = int(data["mark_1"])
        mark_2 = int(data["mark_2"])
        excel_file = load_workbook('python_students.xlsx')
        page = excel_file["Лист1"]
        page.append([new_student, mark_1, mark_2])
        excel_file.save('python_students.xlsx')
        return f"Вы добавили {new_student}!"
    return render_template('create_student.html')


@app.route("/change", methods=["GET", "POST"])
def change():
    if request.method == 'POST':
        data = request.form
        page_row_number = data["row_number"]
        student_name = data["new_name"]
        first = data["mark_1"]
        second = data["mark_2"]
        file_name = "python_students.xlsx"
        excel_file = load_workbook(file_name)
        page = excel_file["Лист1"]
        page[f"A{page_row_number}"] = student_name
        page["B" + page_row_number] = first
        page[f"C{page_row_number}"] = second
        excel_file.save(file_name)
        return f"Строка номер {page_row_number} изменена на '{student_name} - {first} - {second}'"

    return render_template("change.html")


if __name__ == '__main__':
    app.run()
